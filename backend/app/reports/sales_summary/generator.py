"""
Sales Summary — XLSX Generator
================================
Builds a polished, multi-sheet Excel workbook from SalesRow data.

Sheets:
  1. Summary   — pivot-style totals by product
  2. By Month  — full detail table sorted by month → product → region
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.reports.sales_summary.data_source import SalesRow

# ── Palette ────────────────────────────────────────────────────────────────
BRAND_DARK = "1E293B"
BRAND_ACCENT = "6366F1"
HEADER_BG = "6366F1"
HEADER_FG = "FFFFFF"
ALT_ROW_BG = "F1F5F9"
TOTAL_BG = "E0E7FF"
BORDER_COLOR = "CBD5E1"


def _thin_border() -> Border:
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_font() -> Font:
    return Font(name="Calibri", bold=True, color=HEADER_FG, size=11)


def _body_font(bold: bool = False) -> Font:
    return Font(name="Calibri", bold=bold, color=BRAND_DARK, size=10)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=HEADER_BG)


def _alt_fill() -> PatternFill:
    return PatternFill("solid", fgColor=ALT_ROW_BG)


def _total_fill() -> PatternFill:
    return PatternFill("solid", fgColor=TOTAL_BG)


def _apply_header_row(ws, row: int, columns: list[str]) -> None:
    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 1: Summary by Product ────────────────────────────────────────────

def _build_summary_sheet(ws, rows: list[SalesRow]) -> None:
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Sales Summary Report"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color=BRAND_ACCENT)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    subtitle = ws["A2"]
    subtitle.value = "Aggregate totals by product across all regions and months"
    subtitle.font = Font(name="Calibri", size=10, color="64748B", italic=True)
    subtitle.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    headers = ["Product", "Total Units Sold", "Total Revenue (USD)", "Avg Revenue / Unit"]
    _apply_header_row(ws, row=4, columns=headers)
    ws.row_dimensions[4].height = 24

    # Aggregate
    totals: dict[str, dict] = {}
    for r in rows:
        if r.product not in totals:
            totals[r.product] = {"units": 0, "revenue": 0.0}
        totals[r.product]["units"] += r.units_sold
        totals[r.product]["revenue"] += r.revenue_usd

    grand_units = 0
    grand_revenue = 0.0

    for data_row, (product, agg) in enumerate(sorted(totals.items()), start=5):
        fill = _alt_fill() if data_row % 2 == 0 else PatternFill()
        avg = agg["revenue"] / agg["units"] if agg["units"] else 0

        values = [product, agg["units"], agg["revenue"], avg]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=val)
            cell.font = _body_font()
            cell.border = _thin_border()
            cell.fill = fill
            if col_idx in (2,):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            if col_idx in (3, 4):
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")

        grand_units += agg["units"]
        grand_revenue += agg["revenue"]

    # Grand total row
    total_row = 5 + len(totals)
    grand_avg = grand_revenue / grand_units if grand_units else 0
    for col_idx, val in enumerate(
        ["TOTAL", grand_units, grand_revenue, grand_avg], start=1
    ):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font = _body_font(bold=True)
        cell.fill = _total_fill()
        cell.border = _thin_border()
        if col_idx == 2:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        if col_idx in (3, 4):
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")

    _set_col_widths(ws, [24, 20, 24, 24])


# ── Sheet 2: Detail by Month ───────────────────────────────────────────────

def _build_detail_sheet(ws, rows: list[SalesRow]) -> None:
    ws.title = "By Month"
    ws.sheet_view.showGridLines = False

    headers = ["Month", "Product", "Region", "Units Sold", "Revenue (USD)"]
    _apply_header_row(ws, row=1, columns=headers)
    ws.row_dimensions[1].height = 24

    sorted_rows = sorted(rows, key=lambda r: (r.month, r.product, r.region))

    for data_row, r in enumerate(sorted_rows, start=2):
        fill = _alt_fill() if data_row % 2 == 0 else PatternFill()
        values = [r.month, r.product, r.region, r.units_sold, r.revenue_usd]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=val)
            cell.font = _body_font()
            cell.border = _thin_border()
            cell.fill = fill
            if col_idx == 4:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            if col_idx == 5:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    _set_col_widths(ws, [14, 20, 12, 16, 20])


# ── Public entry point ─────────────────────────────────────────────────────

def build_xlsx(data: list[SalesRow], output_path: str) -> None:
    """Build and save the Sales Summary workbook to `output_path`."""
    wb = Workbook()

    # openpyxl creates a default sheet — repurpose it as Summary
    summary_ws = wb.active
    _build_summary_sheet(summary_ws, data)

    detail_ws = wb.create_sheet()
    _build_detail_sheet(detail_ws, data)

    wb.save(output_path)
